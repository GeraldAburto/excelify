import numpy as np
from PIL import Image
from openpyxl import Workbook
from openpyxl.styles import PatternFill
from openpyxl.utils import get_column_letter

row_height = 10
cell_width = 2

# Open image
img = Image.open('cat.jpg')
# Convert image to RGB color matrix
arr = np.array(img)

# Create a new excel sheet
wb = Workbook()
sheet = wb.active

# Convert RGB to Hexadecimal value
def rgb2hex(r,g,b):
    # "#{:02x}{:02x}{:02x}".format(r,g,b) no need '#'
    return "{:02x}{:02x}{:02x}".format(r,g,b)

# Set background color to each cell
for i in range(len(arr)):
    row = i + 1
    sheet.row_dimensions[row].height = row_height
    for j in range(len(arr[i])):
        r = arr[i][j][0]
        g = arr[i][j][1]
        b = arr[i][j][2]
        color = rgb2hex(r, g, b)
        col = j + 1
        sheet.cell(row=row, column=col).fill = PatternFill(fill_type='solid', start_color=color, end_color=color)
        sheet.column_dimensions[get_column_letter(col)].width = cell_width
        
# Save to excel file
wb.save('cat.xlsx')
wb.close()

print('Done!')